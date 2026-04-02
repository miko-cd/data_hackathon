#!/usr/bin/env python3
import csv
import re
import sys
from pathlib import Path
from typing import Optional
import pandas as pd

REGISTER_PATTERN = r'\b[А-ЯӨҮЁ]{2}\d{8}\b'

BAD_WORDS = ["зайл", "хуц", "гөлөг", "новш", "ал", "пизда","секс","сда", "-"]
BAD_WORD_PATTERN = r'\b(' + '|'.join(map(re.escape, BAD_WORDS)) + r')\b'


_ROMAN_VALUES = {'I': 1, 'V': 5, 'X': 10, 'L': 50, 'C': 100, 'D': 500, 'M': 1000}
_ABBR_SUFFIXES = [
    "аас", "ээс",
    "ын", "ийн",
    "д", "т",
    "ыг", "ийг",
    "тай", "тэй",
    "руу", "рүү",
    "аар", "ээр",
]
UNIT_MAP = {
    "kb": "килобайт",
    "mb": "мегабайт",
    "gb": "гигабайт",
    "tb": "терабайт",
    "ms": "миллисекунд",
    "s": "секунд",
    "sec": "секунд",
    "m": "минут",
    "min": "минут",
    "h": "цаг",
    "hr": "цаг",
}

def _pad_if_needed(text: str, start: int, end: int, repl: str) -> str:
    """Add spaces around repl if adjacent chars are Mongolian Cyrillic/Latin letters or digits."""
    pre = text[start - 1] if start > 0 else ''
    post = text[end] if end < len(text) else ''
    if pre and re.match(r'[A-Za-zА-ЯӨҮЁа-яөүё0-9]', pre):
        repl = ' ' + repl
    if post and re.match(r'[A-Za-zА-ЯӨҮЁа-яөүё0-9]', post):
        repl = repl + ' '
    return repl



# ------------------------- brand/operator normalization -------------------------
CANON_REPLACEMENTS = [
    (r'\bapp\b', 'апп'),
    (r'\bunitel\b', 'Юнител'),
    (r'\bg[-\s]?mobile\b', 'Жи мобайл'),
    (r'\bmobicom\b', 'Мобиком'),
    (r'\bskytel\b', 'Скайтел'),
    (r'\bhuawei\b', 'Хуавэй'),
    (r'\bхуа\s*вэй\b', 'Хуавэй'),
    (r'\bхуавей\b', 'Хуавэй'),
    (r'\bхуавэй\b', 'Хуавэй'),
]

def normalize_brands(text: str) -> str:
    for pat, repl in CANON_REPLACEMENTS:
        text = re.sub(pat, repl, text, flags=re.IGNORECASE)
    return text

# ------------------------- numeric pre-normalization -------------------------
def normalize_numeric_formats(text: str) -> str:
    # 100,000 -> 100000 ; 1,234,567 -> 1234567
    def _strip_commas(m: re.Match) -> str:
        return m.group(0).replace(',', '')
    text = re.sub(r'(?<!\d)\d{1,3}(?:,\d{3})+(?!\d)', _strip_commas, text)

    # -25 / −25 / –25 -> "хасах 25" (avoid ranges like 10-20)
    def repl_negative(m: re.Match) -> str:
        return f"хасах {m.group(1)}"
    text = re.sub(r'(?<!\d)[\-\u2212\u2013]\s*(\d+)(?!\d)', repl_negative, text)

    # 800×480 / 800 x 480 -> 800 x 480
    text = re.sub(r'(?<!\d)(\d+)\s*[xX×]\s*(\d+)(?!\d)', r'\1 x \2', text)
    return text

def normalize_units(text: str) -> str:
    unit_keys = sorted(UNIT_MAP.keys(), key=len, reverse=True)
    unit_alt = "|".join(map(re.escape, unit_keys))

    # 100ms -> 100 миллисекунд (keep digits)
    pattern_num_unit = re.compile(rf'(?<!\w)(\d+)\s*({unit_alt})(?!\w)', re.IGNORECASE)

    def repl_num_unit(m: re.Match) -> str:
        n = m.group(1)
        unit = m.group(2).lower()
        word = f"{n} {UNIT_MAP[unit]}"
        return _pad_if_needed(text, m.start(), m.end(), word)

    text = pattern_num_unit.sub(repl_num_unit, text)

    # 3G/4G/5G -> 3 жи (keep digits)
    pattern_g = re.compile(r'(?<!\w)(\d+)\s*G(?!\w)', re.IGNORECASE)

    def repl_g(m: re.Match) -> str:
        n = m.group(1)
        word = f"{n} жи"
        return _pad_if_needed(text, m.start(), m.end(), word)

    text = pattern_g.sub(repl_g, text)
    return re.sub(r'\s+', ' ', text).strip()
def int_to_roman(n: int) -> str:
    if not (0 < n < 4000):
        return ''
    vals = [
        (1000, 'M'), (900, 'CM'),
        (500, 'D'), (400, 'CD'),
        (100, 'C'), (90, 'XC'),
        (50, 'L'), (40, 'XL'),
        (10, 'X'), (9, 'IX'),
        (5, 'V'), (4, 'IV'),
        (1, 'I'),
    ]
    out = []
    for v, sym in vals:
        while n >= v:
            out.append(sym)
            n -= v
    return ''.join(out)

def roman_to_int(s: str) -> int | None:
    s = s.upper()
    if not s or any(ch not in _ROMAN_VALUES for ch in s):
        return None
    total = 0
    prev = 0
    for ch in reversed(s):
        val = _ROMAN_VALUES[ch]
        if val < prev:
            total -= val
        else:
            total += val
            prev = val
    return total if int_to_roman(total) == s else None


def normalize_roman_ordinals(text: str) -> str:
    nouns = ['зуун', 'зууны', 'зуунд', 'зуунаас', 'дайн', 'бүлэг', 'анги', 'зүйл', 'заалт', 'хэсэг', 'боть', 'эгнээ']
    noun_alt = '|'.join(sorted(set(map(re.escape, nouns)), key=len, reverse=True))
    pattern = re.compile(rf'\b([IVXLCDM]+)\s+({noun_alt})\b', flags=re.IGNORECASE)

    back = set('аоуяёАОУЯЁ')      # эр үг
    front = set('эөүиеЭӨҮИЕ')     # эм үг

    def repl(m: re.Match) -> str:
        roman = m.group(1)
        noun = m.group(2)

        n = roman_to_int(roman)
        if n is None:
            return m.group(0)
        n_str = str(n)

        for ch in reversed(n_str):
            if ch in back:
                suffix = 'дугаар'
                break
            if ch in front:
                suffix = 'дүгээр'
                break
        else:
            suffix = 'дугаар'

        return f"{n} {suffix} {noun}"

    return pattern.sub(repl, text)


# mgl abbreviations 
def load_abbreviations_excel(path: str) -> dict[str, str]:
    """
    Excel columns:
      abbr | full
    """
    abbr_map: dict[str, str] = {}
    p = Path(path)
    if not p.exists():
        return abbr_map

    df = pd.read_excel(p)

    cols = {str(c).strip().lower(): c for c in df.columns}
    if "abbr" not in cols or "full" not in cols:
        return abbr_map

    abbr_col = cols["abbr"]
    full_col = cols["full"]

    for _, row in df[[abbr_col, full_col]].iterrows():
        abbr = "" if pd.isna(row[abbr_col]) else str(row[abbr_col]).strip()
        full = "" if pd.isna(row[full_col]) else str(row[full_col]).strip()
        if abbr and full:
            abbr_map[abbr] = full

    return abbr_map

def compile_abbreviation_regex(abbr_map: dict[str, str]) -> Optional[re.Pattern]:
    if not abbr_map:
        return None

    keys = sorted(abbr_map.keys(), key=len, reverse=True)
    abbr_alt = "|".join(map(re.escape, keys))

    dash_alt = r"(?:\-|\u2212|\u2013)"  # -, −, –
    suf_alt = "|".join(map(re.escape, sorted(_ABBR_SUFFIXES, key=len, reverse=True)))

    # group1=abbr, group2=optional suffix
    pat = rf"\b({abbr_alt})(?:{dash_alt}?({suf_alt}))?\b"
    return re.compile(pat)


def expand_abbreviations(text: str, abbr_map: dict[str, str], abbr_re: Optional[re.Pattern]) -> str:
    """
    Expands abbreviations ONLY when the matched token is ALL UPPERCASE.
      - 'ЭНХ' can expand
      - 'Энх' will NOT expand

    Case-sensitive mapping:
      - only expands if the exact key exists in abbr_map
      - no case-insensitive fallback (safe: prevents accidental expansions)
    """
    if not abbr_map or abbr_re is None:
        return text

    def repl(m: re.Match) -> str:
        key = m.group(1)

        # Expand only when fully uppercase
        if not key.isupper():
            return key

        return abbr_map.get(key, key)

    return abbr_re.sub(repl, text)

def compile_abbreviation_regex(abbr_map: dict[str, str]) -> re.Pattern | None:
    if not abbr_map:
        return None
    keys = sorted(abbr_map.keys(), key=len, reverse=True)
    abbr_alt = "|".join(map(re.escape, keys))

    dash_alt = r"(?:\-|\u2212|\u2013)"  # -, −, –
    suf_alt = "|".join(map(re.escape, sorted(_ABBR_SUFFIXES, key=len, reverse=True)))

    # group1 = abbr, group2 = optional suffix
    pat = rf"\b({abbr_alt})(?:{dash_alt}?({suf_alt}))?\b"
    return re.compile(pat)
def _last_vowel_back_or_front(word: str) -> str:
    back = set("аоуяёАОУЯЁ")
    front = set("эөүиеЭӨҮИЕ")
    for ch in reversed(word):
        if ch in back:
            return "back"
        if ch in front:
            return "front"
    return "back"


def _attach_genitive(expanded: str) -> str:
    parts = expanded.split()
    if not parts:
        return expanded
    last = parts[-1]
    kind = _last_vowel_back_or_front(last)
    parts[-1] = last + ("ын" if kind == "back" else "ийн")
    return " ".join(parts)

def _attach_ablative(expanded: str) -> str:
    parts = expanded.split()
    if not parts:
        return expanded
    last = parts[-1]
    kind = _last_vowel_back_or_front(last)
    parts[-1] = last + ("аас" if kind == "back" else "ээс")
    return " ".join(parts)


def _attach_suffix(expanded: str, suffix: str) -> str:
    if not suffix:
        return expanded
    s = suffix.lower()

    if s in ("ын", "ийн"):
        return _attach_genitive(expanded)
    if s in ("аас", "ээс"):
        return _attach_ablative(expanded)

    parts = expanded.split()
    if not parts:
        return expanded + s
    parts[-1] = parts[-1] + s
    return " ".join(parts)

def expand_abbreviations(text: str, abbr_map: dict[str, str], abbr_re: Optional[re.Pattern]) -> str:
    if not abbr_map or abbr_re is None:
        return text

    def repl(m: re.Match) -> str:
        abbr = m.group(1)
        suffix = m.group(2) or ""
        if not abbr.isupper():
            return abbr + (suffix or "")

        expanded = abbr_map.get(abbr, abbr)
        return _attach_suffix(expanded, suffix)

    return abbr_re.sub(repl, text)


def load_currency_country_by_symbol_xlsx(xlsx_path: str) -> dict[str, str]: #pip install openpyxl
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise RuntimeError("Missing dependency: openpyxl. Install: pip install openpyxl") from e
    p = Path(xlsx_path)
    if not p.exists():
        return {}

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header = [str(c.value).strip() if c.value is not None else "" for c in header_row]
    header_l = [h.lower() for h in header]

    def find_col(*names: str) -> int | None:
        for nm in names:
            nm_l = nm.lower()
            if nm_l in header_l:
                return header_l.index(nm_l)
        return None

    country_i = find_col("country", "улс", "country ")
    sym_i = find_col("symbol", "тэмдэг", "symbol ")
    if country_i is None:
        country_i = 0
    if sym_i is None:
        sym_i = 2

    sym_to_country: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2):
        if country_i >= len(row) or sym_i >= len(row):
            continue

        country = row[country_i].value
        sym = row[sym_i].value
        if not country or not sym:
            continue

        country = str(country).strip()
        sym = str(sym).strip()
        if not country or not sym:
            continue
        sym_to_country.setdefault(sym, country)

    return sym_to_country

def compile_currency_symbol_regex(sym_to_country: dict[str, str]) -> re.Pattern | None:
    if not sym_to_country:
        return None
    symbols = sorted(sym_to_country.keys(), key=len, reverse=True)
    sym_alt = "|".join(map(re.escape, symbols))
    return re.compile(rf'({sym_alt})')

def expand_currency_symbols_to_country(text: str, sym_to_country: dict[str, str], sym_re: re.Pattern | None) -> str:
    if not sym_to_country or sym_re is None:
        return text

    def repl(m: re.Match) -> str:
        sym = m.group(1)
        country = sym_to_country.get(sym, sym)

        start, end = m.start(1), m.end(1)
        pre = text[start - 1] if start > 0 else ''
        post = text[end] if end < len(text) else ''

        if post and post.isdigit():
            return f"{country} "
        if pre and pre.isdigit():
            return f" {country}"
        return country

    return sym_re.sub(repl, text)
    
# ------------------------- cleaning pipeline -------------------------
def clean_text(text: str,
               abbr_map: dict[str, str], abbr_re: re.Pattern | None,
               cur_map: dict[str, str], cur_re: re.Pattern | None) -> str:
    text = re.sub(REGISTER_PATTERN, '', text)
    text = re.sub(BAD_WORD_PATTERN, '', text, flags=re.IGNORECASE)

    text = expand_abbreviations(text, abbr_map, abbr_re)

    text = expand_currency_symbols_to_country(text, cur_map, cur_re)
    text = normalize_roman_ordinals(text)

    text = normalize_brands(text)
    text = normalize_numeric_formats(text)
    text = normalize_units(text)

    text = re.sub(r'\s+', ' ', text).strip()
    return text

def main(
    input_path: str,
    output_path: str,
    abbr_xlsx_path: str = "mongol_abb.xlsx",
    currency_xlsx_path: str = "currency_symbols.xlsx",
):
    abbr_map = load_abbreviations_excel(abbr_xlsx_path)
    abbr_re = compile_abbreviation_regex(abbr_map)
    cur_map = load_currency_country_by_symbol_xlsx(currency_xlsx_path)
    cur_re = compile_currency_symbol_regex(cur_map)

    in_path = Path(input_path)
    out_path = Path(output_path)

    if in_path.suffix.lower() == ".txt":
        with open(in_path, encoding="utf-8") as f:
            raw_lines = [line.strip() for line in f if line.strip()]

        cleaned_lines = [clean_text(line, abbr_map, abbr_re, cur_map, cur_re) for line in raw_lines]
        cleaned_lines = [x for x in cleaned_lines if x]

        with open(out_path, "w", encoding="utf-8") as f:
            for line in cleaned_lines:
                f.write(line + "\n")
    else:
        raw_lines = []
        with open(in_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if row:
                    raw_lines.append(row[0])

        cleaned_lines = [clean_text(line, abbr_map, abbr_re, cur_map, cur_re) for line in raw_lines]
        cleaned_lines = [x for x in cleaned_lines if x]

        with open(out_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["text"])
            for line in cleaned_lines:
                w.writerow([line])